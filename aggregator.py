"""Aggregate multiple xlsx files by sheet name.

Each sheet in the output contains data from all source files that have that sheet,
with a source column prepended to identify the origin file.
"""

from __future__ import annotations

import os
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, Optional

import openpyxl

EXCEL_MAX_ROWS = 1_048_576
EXCLUDED_SHEETS = {"Context", "InlineXBRL"}

ProgressCallback = Optional[Callable[[str], None]]


def get_xlsx_files(base_dir: str) -> list[str]:
    xlsx_files = []
    for root, _, files in os.walk(base_dir):
        for file in files:
            if file.endswith(".xlsx"):
                xlsx_files.append(os.path.join(root, file))
    return sorted(xlsx_files)


def _read_sheets_from_file(filepath: str) -> tuple[str, dict[str, list[tuple]], str | None]:
    source_name = Path(filepath).stem
    result = {}
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in EXCLUDED_SHEETS:
                continue
            sheet = wb[sheet_name]
            result[sheet_name] = list(sheet.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return source_name, result, f"{filepath}\t{e}"
    return source_name, result, None


def _make_sheet_name(base: str, part: int) -> str:
    if part == 1:
        return base[:31]
    suffix = f"_{part}"
    return base[: 31 - len(suffix)] + suffix


def aggregate(
    base_dir: str,
    output_path: str,
    workers: int = 8,
    on_progress: ProgressCallback = None,
    cancelled: Callable[[], bool] | None = None,
) -> tuple[bool, list[str]]:
    """Aggregate xlsx files from base_dir into output_path.

    Returns (success, errors).
    """
    if on_progress:
        on_progress(f"Scanning {base_dir} for .xlsx files...")

    xlsx_files = get_xlsx_files(base_dir)
    if not xlsx_files:
        if on_progress:
            on_progress("No .xlsx files found.")
        return False, []

    if on_progress:
        on_progress(f"Found {len(xlsx_files)} files. Reading...")

    sheet_data: dict[str, list[tuple[str, list[tuple]]]] = defaultdict(list)
    errors: list[str] = []
    total = len(xlsx_files)
    done = 0

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_read_sheets_from_file, fp): fp for fp in xlsx_files}

        for future in as_completed(futures):
            if cancelled and cancelled():
                pool.shutdown(wait=False, cancel_futures=True)
                if on_progress:
                    on_progress("Cancelled.")
                return False, errors

            done += 1
            source_name, sheets, error = future.result()
            if error:
                errors.append(error)
                if on_progress:
                    on_progress(f"[{done}/{total}] SKIP: {source_name}")
                continue
            if on_progress:
                on_progress(f"[{done}/{total}] Read: {source_name}")
            for sheet_name, rows in sheets.items():
                sheet_data[sheet_name].append((source_name, rows))

    if cancelled and cancelled():
        return False, errors

    if on_progress:
        on_progress(f"Writing {len(sheet_data)} sheets to {output_path}...")

    wb = openpyxl.Workbook(write_only=True)
    total_rows = 0

    for sheet_name, source_entries in sheet_data.items():
        part = 1
        ws = wb.create_sheet(title=_make_sheet_name(sheet_name, part))
        part_rows = 0

        for source_name, rows in source_entries:
            for row in rows:
                if part_rows >= EXCEL_MAX_ROWS:
                    part += 1
                    ws = wb.create_sheet(title=_make_sheet_name(sheet_name, part))
                    part_rows = 0
                ws.append((source_name,) + (row if row else (None,)))
                total_rows += 1
                part_rows += 1

    wb.save(output_path)

    if on_progress:
        on_progress(f"Done! Saved {output_path} ({total_rows} rows, {len(sheet_data)} sheets)")

    return True, errors
